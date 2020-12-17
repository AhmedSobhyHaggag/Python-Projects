#(1)
#Opening Excel Documents with OpenPyXL
#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#print(type(wb))

#(2)
#Getting Sheets from the Workbook
#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#print(wb.sheetnames) # The workbook's sheets' names.
#sheet = wb['Sheet3'] # Get a sheet from the workbook.
#print(sheet)
#print(type(sheet))
#print(sheet.title) # Get the sheet's title as a string.
#anotherSheet = wb.active # Get the active sheet.
#print(anotherSheet)

#(3)
#Getting Cells from the Sheets
#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1'] # Get a sheet from the workbook.
#sheet['A1'] # Get a cell from the sheet.
#print(sheet['A1'].value) # Get the value from the cell.
#c = sheet['B1'] # Get another cell from the sheet.
#print(c.value)
## Get the row, column, and value from the cell.
#print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
#print('Cell %s is %s' % (c.coordinate, c.value))
#print(sheet['C1'].value)
#print(sheet.cell(row=1, column=2))
#print(sheet.cell(row=1, column=2).value)
#for i in range(1, 8, 2): # Go through every other row:
#     print(i, sheet.cell(row=i, column=2).value)
#print(sheet.max_row) # Get the highest row number.
#print(sheet.max_column) # Get the highest column number.

#(4)
#Converting Between Column Letters and Numbers
#import openpyxl
#from openpyxl.utils import get_column_letter, column_index_from_string
#print(get_column_letter(1)) # Translate column 1 to a letter.
#print(get_column_letter(2))
#print(get_column_letter(27))
#print(get_column_letter(900))
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1']
#print(get_column_letter(sheet.max_column))
#print(column_index_from_string('A')) # Get A's number.
#print(column_index_from_string('AA'))

#(5)
#Getting Rows and Columns from the Sheets
#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1']
#print(tuple(sheet['A1':'C3'])) # Get all cells from A1 to C3
#for rowOfCellObjects in sheet['A1':'C3']:
#    for cellObj in rowOfCellObjects:
#        print(cellObj.coordinate, cellObj.value)
#    print('--- END OF ROW ---')

#(6)
#To access the values of cells in a particular row or column, you can also use a Worksheet object’s rows and columns attribute
#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb.active
#print(list(sheet.columns)[1]) # Get second column's cells.
#for cellObj in list(sheet.columns)[1]:
#    print(cellObj.value)

#(7)
#Project: Reading Data from a Spreadsheet
#Step 1: Read the Spreadsheet Data
 #! python3
   # readCensusExcel.py - Tabulates population and number of census tracts for
   # each county.
#import openpyxl, pprint
#print('Opening workbook...')
#wb = openpyxl.load_workbook('censuspopdata.xlsx')
#sheet = wb['Population by Census Tract']
#countyData = {}
## TODO: Fill in countyData with each county's population and tracts.
#print('Reading rows...')
#for row in range(2, sheet.max_row + 1):
## Each row in the spreadsheet has data for one census tract.
#    state  = sheet['B' + str(row)].value
#    county = sheet['C' + str(row)].value
#    pop    = sheet['D' + str(row)].value
#    # Make sure the key for this state exists.
#    countyData.setdefault(state, {})
#    # Make sure the key for this county in this state exists.
#    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
#    # Each row represents one census tract, so increment by one.
#    countyData[state][county]['tracts'] += 1
#    # Increase the county pop by the pop in this census tract.
#    countyData[state][county]['pop'] += int(pop)
##Step 3: Write the Results to a File
## Open a new text file and write the contents of countyData to it.
#print('Writing results to a python file...')
#resultFile = open('census2010.py', 'w')
#resultFile.write('allData = ' + pprint.pformat(countyData))
#resultFile.close()
#print('Done.')
##Step 4: Whenever you need the county data, you can just run import census2010
#import os
#import census2010
#print(census2010.allData['AK']['Anchorage'])
#anchoragePop = census2010.allData['AK']['Anchorage']['pop']
#print('The 2010 population of Anchorage was ' + str(anchoragePop))

#(8)
#Writing Excel Documents
#Creating and Saving Excel Documents
#import openpyxl
#wb = openpyxl.Workbook() # Create a blank workbook.
#print(wb.sheetnames) # It starts with one sheet.
#sheet = wb.active
#print(sheet.title)
#sheet.title = 'Spam Bacon Eggs Sheet' # Change title.
#print(wb.sheetnames)

#(9)
#the spreadsheet file will not be saved until you call the save() workbook method
#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb.active
#sheet.title = 'Spam Spam Spam'
#wb.save('example_copy.xlsx') # Save the workbook.

#(10)
#Creating and Removing Sheets
#import openpyxl
#wb = openpyxl.Workbook()
#print(wb.sheetnames)
#print(wb.create_sheet()) # Add a new sheet.
#print(wb.sheetnames)
## Create a new sheet at index 0.
#print(wb.create_sheet(index=0, title='First Sheet'))
#print(wb.sheetnames)
#print(wb.create_sheet(index=2, title='Middle Sheet'))
#print(wb.sheetnames)
#del wb['Middle Sheet']
#del wb['Sheet1']
#print(wb.sheetnames)
#wb.save('another_example1.xlsx')

#(11)
#Writing Values to Cells
#import openpyxl
#wb = openpyxl.load_workbook('another_example1.xlsx')
#sheet = wb['Sheet']
#sheet['A1'] = 'Hello, world!'   Edit the cell's value.
#print(sheet['A1'].value)
#wb.save('another_example1.xlsx')

#(12)
#Project: Updating a Spreadsheet
#Step 1: Set Up a Data Structure with the Update Information
#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.
#import openpyxl
#wb = openpyxl.load_workbook('produceSales.xlsx')
#sheet = wb['Sheet']
## The produce types and their updated prices
#PRICE_UPDATES = {'Garlic': 3.07,'Celery': 1.19,'Lemon': 1.27}
##Step 2: Check All Rows and Update Incorrect Prices
## Loop through the rows and update the prices.
#for rowNum in range(2, sheet.max_row):    # skip the first row
#    produceName = sheet.cell(row=rowNum, column=1).value
#    if produceName in PRICE_UPDATES:
#       sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
#wb.save('updatedProduceSales.xlsx')

#(13)
#Setting the Font Style of Cells
#import openpyxl
#from openpyxl.styles import Font
#wb = openpyxl.Workbook()
#sheet = wb['Sheet']
#italic24Font = Font(size=24, italic=True) # Create a font.
#sheet['A1'].font = italic24Font # Apply the font to A1.
#sheet['A1'] = 'Hello, world!'
#wb.save('styles.xlsx')

#(14)
#Font Objects
#import openpyxl
#from openpyxl.styles import Font
#wb = openpyxl.Workbook()
#sheet = wb['Sheet']
#fontObj1 = Font(name='Times New Roman', bold=True)
#sheet['A1'].font = fontObj1
#sheet['A1'] = 'Bold Times New Roman'
#fontObj2 = Font(size=24, italic=True)
#sheet['B3'].font = fontObj2
#sheet['B3'] = '24 pt Italic'
#wb.save('styles.xlsx')

#(15)
#Formulas
#import openpyxl
#wb = openpyxl.Workbook()
#sheet = wb.active
#sheet['A1'] = 200
#sheet['A2'] = 500
#sheet['A3'] = '=AVERAGE(A1:A2)' # Set the formula.
#wb.save('styles.xlsx')

#(16)
#Adjusting Rows and Columns
#import openpyxl
#wb = openpyxl.Workbook()
#sheet = wb.active
#sheet['A1'] = 'Tall row'
#sheet['B2'] = 'Wide column'
## Set the height and width:
#sheet.row_dimensions[1].height = 70
#sheet.column_dimensions['B'].width = 20
#wb.save('styles.xlsx')

#(17)
#Merging and Unmerging Cells
#import openpyxl
#wb = openpyxl.Workbook()
#sheet = wb.active
##Merging
#sheet.merge_cells('A1:D3') # Merge all these cells.
#sheet['A1'] = 'Twelve cells merged together.'
#sheet.merge_cells('C5:D5') # Merge these two cells.
#sheet['C5'] = 'Two merged cells.'
#wb.save('styles.xlsx')
##UnMerging
#import openpyxl
#wb = openpyxl.load_workbook('styles.xlsx')
#sheet = wb.active
#sheet.unmerge_cells('A1:D3') # Split these cells up.
#sheet.unmerge_cells('C5:D5')
#wb.save('styles.xlsx')

#(18)
#Freezing Panes
#import openpyxl
#wb = openpyxl.load_workbook('produceSales.xlsx')
#sheet = wb.active
#sheet.freeze_panes = 'A2' # Freeze the rows above A2.
#wb.save('updatedProduceSales.xlsx')

#(19)
#Charts
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11): # create some data in column A
    sheet['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1,max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.PieChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('styles.xlsx')